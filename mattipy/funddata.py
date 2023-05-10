import pandas as pd
from openpyxl import load_workbook

# Step 1: Read the Excel file and get the font properties
file_path = '/Users/matthias/Desktop/Fjarmalaverkfr2023/VerkfraediX/Symmetria/Funddata/funddatafun.xlsx'
sheet_names = ['1999_sam']
workbook = load_workbook(file_path)

#sheet_names = ['2018_sam', '2017_sam', '2016_sam', '2015_sam', '2014_sam', '2013_sam', '2012_sam', '2011_sam', '2010_sam', '2009_sam', '2008_sam', '2007_sam', '2006_sam', '2005_sam', '2004_sam', '2003_sam', '2002_sam', '2001_sam', '2000_sam', '1999_sam', '1998_sam', '1997_sam', '2018_ser', '2017_ser', '2016_ser', '2015_ser', '2014_ser', '2013_ser', '2012_ser', '2011_ser', '2010_ser', '2009_ser', '2008_ser', '2007_ser', '2006_ser', '2005_ser', '2004_ser', '2003_ser', '2002_ser', '2001_ser', '2000_ser', '1999_ser', '1998_ser', '1997_ser']

for sheet_name in sheet_names:
    print(sheet_name)
    worksheet = workbook[sheet_name]

    cells = []
    for row in worksheet.iter_rows():
        first_cell = row[0]  # Process the first cell in each row
        second_cell = row[1]  # Process the second cell in each row
        cell_data = {
            'value': first_cell.value,
            'font_size': first_cell.font.sz,
            'bold': first_cell.font.b,
            'italic': first_cell.font.i,
            'yes_column_value': second_cell.value, # Store the value of the second cell
            'values': [cell.value if cell.value is not None and cell.value not in ['*', '**', '***'] else "NULL" for cell in row[7:]]
        }
        cells.append(cell_data)

    # Step 2: Process the data based on font properties
    data = []
    current_attributes = ['', '', '', '']

    for cell_data in cells:
        value, font_size, bold, italic, yes_column_value = cell_data['value'], cell_data['font_size'], cell_data['bold'], cell_data['italic'], cell_data['yes_column_value']
        
        if value is None:
            continue

        if italic:
            value = '(Samtals) ' + value

        if font_size == 14.0 and bold:  # Attribute 1
            current_attributes[0] = value
            current_attributes[1] = ''
            current_attributes[2] = ''
            current_attributes[3] = ''
        elif font_size == 8.0 and bold:  # Attribute 2
            current_attributes[1] = value
            current_attributes[2] = ''
            current_attributes[3] = ''
        elif font_size == 8.0 and not bold:  # Attribute 3
            current_attributes[2] = value
            current_attributes[3] = ''
        elif font_size == 6.0 and not bold:  # Attribute 4
            current_attributes[3] = value

        # Check for 'YES' in the second column before appending data
        if yes_column_value == 'YES':
            row_dict = {
                'Attribute 1': current_attributes[0],
                'Attribute 2': current_attributes[1],
                'Attribute 3': current_attributes[2],
                'Attribute 4': current_attributes[3],
                '': '',
            }
            
            for i, val in enumerate(cell_data['values']):
                row_dict[f'Value {i+1}'] = val
            
            data.append(row_dict)

    # Step 3: Create a new DataFrame and export to a new Excel file

    new_df = pd.DataFrame(data)

    output_file_path = '/Users/matthias/Desktop/Fjarmalaverkfr2023/VerkfraediX/Symmetria/Funddata/'+sheet_name+'_output'+'.xlsx'
    with pd.ExcelWriter(output_file_path) as writer:
        new_df.to_excel(writer, index=False, startrow=2)